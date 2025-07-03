#!/usr/bin/env python3
"""
AGREGAR CANALES MANUALES AL SISTEMA
Procesa una lista de IDs de canales usando los mismos criterios del script principal
"""

import os
import sys
import json
import time
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

# Importar las clases del script principal
# Asumiendo que el script principal está en el mismo directorio
try:
    from main import (
        Config, Logger, OptimizedYouTubeClient, ChannelProcessor, 
        DataManager, StreamerData, ArgentineDetector, LiveContentDetector, 
        ContentCategorizer, QuotaTracker, APICache
    )
except ImportError:
    print("❌ Error: No se puede importar desde main.py")
    print("Asegúrate de que este script esté en el mismo directorio que main.py")
    sys.exit(1)

class ManualChannelProcessor:
    """Procesador específico para canales añadidos manualmente"""
    
    def __init__(self, channel_ids: List[str]):
        Config.setup_directories()
        
        self.channel_ids = channel_ids
        self.logger_system = Logger()
        self.logger = logging.getLogger('StreamingArgentina')
        
        # Inicializar componentes usando las mismas clases del script principal
        self.youtube = OptimizedYouTubeClient()
        self.processor = ChannelProcessor(self.youtube, self.logger_system)
        
        self.processed_count = 0
        self.added_count = 0
        self.rejected_count = 0
        
    def process_all_channels(self):
        """Procesar todos los canales de la lista"""
        
        self.logger.info("="*80)
        self.logger.info("📝 PROCESANDO CANALES MANUALES")
        self.logger.info(f"📅 Fecha: {datetime.now():%Y-%m-%d %H:%M:%S}")
        self.logger.info(f"📊 Total de canales: {len(self.channel_ids)}")
        self.logger.info(f"🔋 Cuota disponible: {self.youtube.quota_tracker.get_remaining():,}")
        self.logger.info("="*80)
        
        # Procesar cada canal
        for i, channel_id in enumerate(self.channel_ids, 1):
            if not self.youtube.can_continue():
                self.logger.warning("⚠️ Deteniendo procesamiento - Límites alcanzados")
                break
            
            self.logger.info(f"\n🔍 Procesando canal {i}/{len(self.channel_ids)}: {channel_id}")
            
            try:
                result = self.process_single_channel(channel_id)
                
                if result:
                    self.added_count += 1
                    self.logger.info(f"✅ Canal añadido: {result.nombre_canal}")
                else:
                    self.rejected_count += 1
                
                self.processed_count += 1
                
                # Log de progreso cada 10 canales
                if i % 10 == 0:
                    self.logger_system.log_quota_status(
                        self.youtube.quota_tracker.used_today,
                        Config.MAX_DAILY_QUOTA
                    )
                    self.logger.info(f"📊 Progreso: {i}/{len(self.channel_ids)} | "
                                   f"Añadidos: {self.added_count} | "
                                   f"Rechazados: {self.rejected_count}")
                
                # Pausa para no saturar la API
                time.sleep(0.5)
                
            except Exception as e:
                self.logger.error(f"❌ Error procesando canal {channel_id}: {e}")
                self.rejected_count += 1
                continue
        
        # Guardar estado final
        self.processor.data_manager.save_state()
        self.youtube.cache.save_cache()
        self.youtube.quota_tracker.save_quota()
        
        # Mostrar resumen
        self.print_final_summary()
        
        return {
            'processed': self.processed_count,
            'added': self.added_count,
            'rejected': self.rejected_count
        }
    
    def process_single_channel(self, channel_id: str) -> Optional[StreamerData]:
        """Procesar un canal individual usando los mismos criterios del script principal"""
        
        # Verificar si ya fue procesado
        if self.processor.data_manager.is_processed(channel_id):
            self.logger.info(f"   ⏭️ Canal ya procesado anteriormente")
            return None
        
        # Crear snippet fake para usar el método del procesador principal
        fake_snippet = {
            'id': {'channelId': channel_id}
        }
        
        # Usar el procesador principal que ya tiene todos los filtros
        result = self.processor.process_channel(fake_snippet)
        
        return result
    
    def print_final_summary(self):
        """Mostrar resumen final"""
        
        print("\n" + "="*80)
        print("📊 RESUMEN FINAL - CANALES MANUALES")
        print("="*80)
        print(f"📊 Total procesados: {self.processed_count}")
        print(f"✅ Canales añadidos: {self.added_count}")
        print(f"❌ Canales rechazados: {self.rejected_count}")
        print(f"💾 Datos guardados en: {Config.STREAMERS_CSV}")
        
        if self.added_count > 0:
            success_rate = (self.added_count / self.processed_count) * 100
            print(f"📈 Tasa de éxito: {success_rate:.1f}%")
        
        print("="*80)

def load_channel_ids_from_file(file_path: str) -> List[str]:
    """Cargar IDs de canales desde un archivo"""
    
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"No se encuentra el archivo: {file_path}")
    
    channel_ids = []
    
    if file_path.suffix.lower() == '.json':
        # Archivo JSON
        with open(file_path, 'r') as f:
            data = json.load(f)
            if isinstance(data, list):
                channel_ids = data
            elif isinstance(data, dict) and 'channels' in data:
                channel_ids = data['channels']
            else:
                raise ValueError("Formato JSON no reconocido")
    
    elif file_path.suffix.lower() == '.txt':
        # Archivo de texto, un ID por línea
        with open(file_path, 'r') as f:
            channel_ids = [line.strip() for line in f if line.strip()]
    
    elif file_path.suffix.lower() == '.csv':
        # Archivo CSV
        import csv
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            
            # Intentar detectar automáticamente la columna con IDs
            first_row = next(reader, None)
            if not first_row:
                raise ValueError("El archivo CSV está vacío")
            
            # Buscar columna que contenga IDs o URLs de YouTube
            id_column = None
            for i, header in enumerate(first_row):
                if any(keyword in header.lower() for keyword in ['id', 'channel', 'canal', 'url', 'link']):
                    id_column = i
                    break
            
            # Si no encuentra header específico, usar primera columna
            if id_column is None:
                print("⚠️ No se detectó columna de IDs, usando primera columna")
                id_column = 0
            
            print(f"📊 Usando columna {id_column + 1}: '{first_row[id_column] if id_column < len(first_row) else 'N/A'}'")
            
            # Leer todas las filas
            f.seek(0)  # Volver al inicio
            reader = csv.reader(f)
            next(reader)  # Saltar header
            
            for row in reader:
                if len(row) > id_column and row[id_column].strip():
                    channel_ids.append(row[id_column].strip())
    
    elif file_path.suffix.lower() == '.xlsx':
        # Archivo Excel (requiere openpyxl)
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            print(f"📊 Leyendo hoja: {sheet.title}")
            
            # Buscar columna con IDs
            header_row = 1
            id_column = None
            
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value and any(keyword in str(cell_value).lower() for keyword in ['id', 'channel', 'canal', 'url', 'link']):
                    id_column = col
                    break
            
            if id_column is None:
                print("⚠️ No se detectó columna de IDs, usando primera columna")
                id_column = 1
            
            # Leer datos
            for row in range(header_row + 1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=id_column).value
                if cell_value and str(cell_value).strip():
                    channel_ids.append(str(cell_value).strip())
            
        except ImportError:
            raise ValueError("Para leer archivos Excel, instala: pip install openpyxl")
    
    else:
        raise ValueError("Formato de archivo no soportado. Usa .json, .txt, .csv o .xlsx")
    
    # Limpiar y validar IDs
    cleaned_ids = []
    for channel_id in channel_ids:
        channel_id = channel_id.strip()
        
        # Extraer ID si es una URL
        if 'youtube.com/channel/' in channel_id:
            channel_id = channel_id.split('/channel/')[-1].split('?')[0]
        elif 'youtube.com/c/' in channel_id or 'youtube.com/@' in channel_id:
            print(f"⚠️ Advertencia: {channel_id} parece ser un handle/username, no un ID de canal")
            continue
        
        # Validar formato de ID
        if channel_id.startswith('UC') and len(channel_id) == 24:
            cleaned_ids.append(channel_id)
        else:
            print(f"⚠️ ID inválido ignorado: {channel_id}")
    
    return cleaned_ids

def main():
    """Función principal"""
    
    try:
        # Verificar configuración
        if not Config.YOUTUBE_API_KEY:
            print("❌ ERROR: No hay API key configurada")
            print("Configura tu API key:")
            print("  export YOUTUBE_API_KEY='tu_api_key'")
            return
        
        # Buscar archivo con IDs de canales
        possible_files = [
            'CANALES ARGENTINOS FINALES.xlsx',  # Tu archivo Excel
            'channel_ids.xlsx',
            'canales.xlsx',
            'channel_ids.csv',
            'canales.csv',
            'channel_ids.json',
            'canales.json', 
            'channel_ids.txt',
            'canales.txt',
            'manual_channels.json',
            'manual_channels.txt'
        ]
        
        channel_ids = []
        used_file = None
        
        # Intentar cargar desde archivos existentes
        for file_name in possible_files:
            if Path(file_name).exists():
                try:
                    channel_ids = load_channel_ids_from_file(file_name)
                    used_file = file_name
                    break
                except Exception as e:
                    print(f"⚠️ Error cargando {file_name}: {e}")
                    continue
        
        # Si no hay archivo, crear lista manual
        if not channel_ids:
            print("📝 No se encontraron archivos con IDs. Usando lista manual...")
            
            # AQUÍ PUEDES PONER TUS 189 IDs DE CANALES
            channel_ids = [
                # Ejemplo de IDs - reemplaza con tu lista real
                'UCxxxxxxxxxxxxxxxxxxxx1',
                'UCxxxxxxxxxxxxxxxxxxxx2',
                'UCxxxxxxxxxxxxxxxxxxxx3',
                # ... añade tus 189 IDs aquí
            ]
            
            # O crear un archivo ejemplo
            if not channel_ids or channel_ids[0].startswith('UCxxxxx'):
                sample_file = Path('channel_ids.json')
                sample_data = {
                    "description": "Lista de IDs de canales de YouTube argentinos",
                    "channels": [
                        "UCxxxxxxxxxxxxxxxxxxxxxxx1",
                        "UCxxxxxxxxxxxxxxxxxxxxxxx2", 
                        "UCxxxxxxxxxxxxxxxxxxxxxxx3"
                    ]
                }
                
                with open(sample_file, 'w') as f:
                    json.dump(sample_data, f, indent=2)
                
                print(f"📁 Archivo ejemplo creado: {sample_file}")
                print("Edita este archivo con tus IDs de canales y ejecuta el script nuevamente.")
                return
        
        if not channel_ids:
            print("❌ No hay IDs de canales para procesar")
            return
        
        print(f"📋 Cargados {len(channel_ids)} IDs de canales desde: {used_file or 'lista manual'}")
        
        # Confirmar antes de procesar
        print(f"\n🔍 Se procesarán {len(channel_ids)} canales usando los mismos criterios del script principal:")
        print(f"   - Mínimo {Config.MIN_SUBSCRIBERS} suscriptores")
        print(f"   - Mínimo {Config.MIN_CERTAINTY_ARGENTINA}% certeza de ser argentino")
        print(f"   - Mínimo {Config.MIN_CERTAINTY_STREAMING}% certeza de hacer streaming")
        
        response = input("\n¿Continuar? (y/N): ").strip().lower()
        if response not in ['y', 'yes', 's', 'si', 'sí']:
            print("❌ Cancelado por el usuario")
            return
        
        # Procesar canales
        processor = ManualChannelProcessor(channel_ids)
        results = processor.process_all_channels()
        
        print(f"\n✅ Procesamiento completado:")
        print(f"   - Procesados: {results['processed']}")
        print(f"   - Añadidos: {results['added']}")
        print(f"   - Rechazados: {results['rejected']}")
        
    except KeyboardInterrupt:
        print("\n⚠️ Procesamiento interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error crítico: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
